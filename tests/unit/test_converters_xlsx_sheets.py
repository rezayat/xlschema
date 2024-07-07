import openpyxl
import pytest

from conftest import xlsx
from xlschema.config import Config

ROW_OFFSET = len(Config.METAFIELDS)
COL_OFFSET = 1

@pytest.fixture(scope="module")
def sheet():
    in_file = xlsx('test-simple')
    wb = openpyxl.load_workbook(in_file)
    sheet = wb.worksheets[0]
    yield sheet
    wb.close()

def test_sheet_iter_rows(sheet):
    results = []
    # for row in sheet.iter_rows(row_offset=ROW_OFFSET):
    for row in sheet.iter_rows(min_row=ROW_OFFSET+1):    
        results.append(list(r.value for r in row))
    assert results[0][COL_OFFSET:] == [1, 'horse', 'horse1']

def test_sheet_rows(sheet):
    results = []
    for row in sheet.rows:
        results.append(list(r.value for r in row))
    assert results[ROW_OFFSET][COL_OFFSET:] == [1, 'horse', 'horse1']

def test_sheet(sheet):
    from xlschema.readers.xlsx.sheets import XlSheet
    sheet = XlSheet(sheet)
    assert repr(sheet) == "<XlSheet 'data'>"

# def test_no_data_empty():
#     in_file = xlsx('test-no-data-empty')
#     wb = openpyxl.load_workbook(in_file)
#     sheet = wb.worksheets[0]
#     from xlschema.readers.xlsx.sheets import NoDataSheet
#     sheet = NoDataSheet(sheet)
#     sheet.parse()
