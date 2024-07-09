"""Excel-oriented Writer classes.

Writer classes to write Excel files::

    SchemaWriter
        ExcelWriter
"""
import openpyxl

from ..config import Config, register
from ..readers.xlsx.mixins import ToExcelMixin
from .abstract import SchemaWriter


@register
class ExcelWriter(SchemaWriter, ToExcelMixin):
    """Convert parsed abstract model into machine readable xl format."""

    file_suffix = 'xlsx'
    method = 'validation'

    def _write_sheet_metadata(self, model, sheet):
        """Write metadata to sheet."""
        self.log.debug("writing %s.metadata", model.name)
        for i, label in enumerate(Config.METAFIELDS):
            cell = sheet.cell(row=i + 1, column=1)
            self.style_metadata(cell)
            cell.value = label

    def _write_sheet_header(self, model, sheet):
        """Write header to sheet."""
        self.log.debug("writing %s.header:", model.name)
        for i, field in enumerate(model.fields, 1):
            i += 1  # to offset column by 1
            self.log.debug("\tfield: %s %s", field.type, field.name)
            for j, name in enumerate(Config.METAFIELDS):
                j += 1
                cell = sheet.cell(row=j, column=i)
                cell.value = getattr(field, name)
                if name in ['name']:
                    self.style_fieldnames(cell)
                else:
                    self.style_table(cell)

    def _write_sheet_header_validation(self, model, sheet):
        """Write header validation to sheet."""
        self.log.debug("adding %s.header_validation:", model.name)
        col_end = self.xl_column(len(model.fields) + 1)

        def _row_range(row_n):
            """Help row formater."""
            return 'B{}:{}{}'.format(row_n, col_end, row_n)

        # set 1: required (1|None)
        self.valid_list(
            sheet, _row_range(1), self.config.ACTIONS, allow_blank=True)
        # set 6: required (1|None)
        self.valid_list(sheet, _row_range(6), [1, 0], allow_blank=True)
        # set 7: index validation (pk|fk|pfk|sk|None)
        self.valid_list(
            sheet, _row_range(7), ['pk', 'fk', 'pfk', 'sk'], allow_blank=True)
        # set 8: length (whole number)
        self.valid_int(sheet, _row_range(8))
        # set 9: type (list)
        self.valid_list(
            sheet,
            _row_range(9),
            self.field_class.VALID_TYPES,
            allow_blank=False)

    def _write_sheet_values(self, model, sheet):
        """Write values to sheet."""
        if model.data:
            self.log.debug("writing %s.values:", model.name)
            for i, row in enumerate(model.data, 1):
                self.log.debug("\trow: %s", row)
                for j, field in enumerate(row, 1):
                    j += 1  # to offset column by 1
                    cell = sheet.cell(row=i + self.n_args, column=j)
                    cell.value = field
                    self.style_table(cell)
        else:
            self.log.warning("no data, writing styles only")
            for i in range(1, self.n_args):  # for symmetry
                for j in range(1, len(model.fields) + 1):
                    j += 1  # to offset column by 1
                    cell = sheet.cell(row=i + self.n_args, column=j)
                    self.style_table(cell)

    def _write_sheet_data_validation(self, model, sheet):
        """Write data validation to sheet."""
        self.log.debug("adding %s.data_validation:", model.name)
        if not model.data:
            nrows_data = self.n_args  # set number of empty rows for symmetry
            self.log.warning("no data, writing validation rules only")
        else:
            nrows_data = len(model.data)

        for i, field in enumerate(model.fields, 1):
            col = self.xl_column(i + 1)
            row_start = self.n_args + 1
            row_end = row_start + nrows_data - 1
            ref = '{}{}'.format(col, row_start)
            col_range = '{}:{}{}'.format(ref, col, row_end)
            if field.is_enum:
                try:
                    enum = self.schema.enums[field.name]
                except KeyError:
                    self.log.critical(
                        "Enum not specified for %s.%s", model.name, field.name)
                    raise
                self.log.debug("\tenum %s -> %s", field.name,
                               list(enum.keys()))
                self.valid_list(sheet, col_range,
                                list(enum.keys()), not field.required)

            elif field.type == 'int':
                self.log.debug("\t%s -> int validation", field.name)
                self.valid_int(sheet, col_range)

            elif field.type in ['dec', 'float', 'double', 'numeric']:
                self.log.debug("\t%s -> decimal validation", field.name)
                self.valid_dec(sheet, col_range)

            elif field.length and (field.type in ['str', 'txt']):
                self.log.debug(
                    "\t%s -> str length validation", field.name)
                self.valid_str_len(sheet, col_range, field.length)

            else:
                continue  # pragma: no cover

    def _write_workbook_enums(self, workbook):
        """Write enums to workbook."""
        if Config.ENUMS_SHEET not in workbook.sheetnames:
            sheet = workbook.create_sheet()
            sheet.title = Config.ENUMS_SHEET
            sheet.sheet_view.showGridLines = False
            nrows = 0
            for name in sorted(self.schema.enums.keys()):
                nrows += 1
                cell = sheet.cell(row=nrows, column=1)
                cell.value = name
                self.style_enum(cell)
                sheet.merge_cells('A{row}:B{row}'.format(row=nrows))
                enum = self.schema.enums[name]
                for key, val in enum.data:
                    nrows += 1
                    cell_key = sheet.cell(row=nrows, column=1)
                    cell_key.value = key
                    self.style_normal(cell_key)
                    cell_val = sheet.cell(row=nrows, column=2)
                    cell_val.value = val
                    self.style_left(cell_val)
                nrows += 1

    def write(self, to_path=None):
        """Write xlsx file from models."""
        if to_path:
            path = to_path
        else:
            path = self.path
        self.log.debug('starting validation pass')
        workbook = openpyxl.Workbook()
        sheet_gen = self.gen_sheets(workbook)
        self.log.debug("writing: %s", path)

        for model in self.schema.models:
            name = model.name
            self.log.debug("writing model: %s", name)
            sheet = next(sheet_gen)
            sheet.title = name

            # cosmetic adjustments
            sheet.sheet_view.showGridLines = False
            sheet.freeze_panes = 'B1'
            self.autowidth(model, sheet)

            # write metadata
            self._write_sheet_metadata(model, sheet)

            # write header
            self._write_sheet_header(model, sheet)

            # writer header validation
            self._write_sheet_header_validation(model, sheet)

            # write values
            self._write_sheet_values(model, sheet)

            # set validation
            self._write_sheet_data_validation(model, sheet)

        # create ENUMs
        self._write_workbook_enums(workbook)

        workbook.save(path)
