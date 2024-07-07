"""Mixin classes to support xl-specific styles and validation."""

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, colors
from openpyxl.worksheet.datavalidation import DataValidation

from ...common.list import List

# ----------------------------------------------------------
# XLSX MIXINS
# ----------------------------------------------------------


class ToExcelMixin(object):
    """Mixin class to hold properties and useful methods in writing."""

    XL_MAXROWS = 1048576
    XL_FONT = 'Calibri'
    XL_FONT_SIZE = 9
    XL_BORDER_COLOR = colors.COLOR_INDEX[22]
    XL_BG_COLOR = '00EAEAEA'
    XL_FONT_NORMAL = Font(name=XL_FONT, size=XL_FONT_SIZE)
    XL_FONT_BOLD = Font(name=XL_FONT, size=XL_FONT_SIZE, bold=True)
    XL_FONT_ITALIC = Font(name=XL_FONT, size=XL_FONT_SIZE, italic=True)
    XL_ALIGN_LEFT = Alignment(horizontal='left')
    XL_ALIGN_CENTER = Alignment(horizontal='center')
    XL_ALIGN_RIGHT = Alignment(horizontal='right', indent=1)

    XL_BORDER = Border(
        left=Side(style='thin', color=XL_BORDER_COLOR),
        right=Side(style='thin', color=XL_BORDER_COLOR),
        top=Side(style='thin', color=XL_BORDER_COLOR),
        bottom=Side(style='thin', color=XL_BORDER_COLOR))
    XL_FILL = PatternFill(
        fill_type='solid', start_color=XL_BG_COLOR, end_color=XL_BG_COLOR)
    XL_VALIDATION_ERROR_TITLE = 'Invalid Entry'
    XL_COLUMN_WIDTH_METHOD = 'max'  # can be max|median|int|None

    # pylint: disable=no-self-use
    def gen_sheets(self, workbook, start=0):
        """Generates workbook sheets.

        Usage::

            sheet_gen = self.gen_sheets(workbook)
            ...
            for model in models:
                sheet = next(sheet_gen)
        """
        i = start
        while True:
            if i == start:
                yield workbook.active
            else:
                yield workbook.create_sheet()
            i += 1

    def autowidth(self, model, sheet):
        """Set columns to width of field names."""
        column_widths = []
        field_lengths = [len(field.name) for field in model.fields]
        method = self.XL_COLUMN_WIDTH_METHOD
        # we default to max
        # column_widths = [max(field_length) for field_length in field_lengths]
        for length in field_lengths:
            if method == 'max':
                column_widths.append(max(field_lengths))
            elif method == 'median':
                column_widths.append(List.median(field_lengths))
            elif isinstance(method, int):
                column_widths.append(method)
            else:
                column_widths.append(length)

        for i, column_width in enumerate(column_widths):
            i += 1
            sheet.column_dimensions[self.xl_column(i + 1)].width = column_width

    def xl_column(self, coln):
        """Converts a column number to alphabetical excel column label."""
        quot, rem = divmod(coln - 1, 26)
        return self.xl_column(quot) + chr(rem + ord('A')) if coln != 0 else ''

    def style_normal(self, cell):
        """Set base style."""
        cell.font = self.XL_FONT_NORMAL
        cell.alignment = self.XL_ALIGN_CENTER

    def style_left(self, cell):
        """Set base style."""
        cell.font = self.XL_FONT_NORMAL
        cell.alignment = self.XL_ALIGN_LEFT

    def style_table(self, cell):
        """Set style of table cells."""
        self.style_normal(cell)
        cell.border = self.XL_BORDER

    def style_bold_filled(self, cell):
        """Set style of bold and grey-filled cells."""
        cell.font = self.XL_FONT_BOLD
        cell.alignment = self.XL_ALIGN_CENTER
        cell.fill = self.XL_FILL

    def style_enum(self, cell):
        """Set style of enum titles."""
        self.style_bold_filled(cell)

    def style_fieldnames(self, cell):
        """Set style of fieldname cells."""
        self.style_bold_filled(cell)
        cell.border = self.XL_BORDER

    def style_metadata(self, cell):
        """Set style of metadata cells."""
        cell.font = self.XL_FONT_BOLD
        cell.alignment = self.XL_ALIGN_RIGHT

    # def style_index(self, cell):
    #     """Set style of index cells
    #     """
    #     cell.font = self.XL_FONT_ITALIC
    #     cell.alignment = self.XL_ALIGN_CENTER
    #     cell.border = self.XL_BORDER

    def valid_list(self, sheet, xlrange, keys, allow_blank=True):
        """Set sheet range to list validation."""
        formula = '"{}"'.format(','.join([str(k) for k in keys]))
        dvrule = DataValidation(
            type="list", formula1=formula, allow_blank=allow_blank)
        dvrule.errorTitle = self.XL_VALIDATION_ERROR_TITLE
        dvrule.error = 'Your entry is not in the list.'
        sheet.add_data_validation(dvrule)
        dvrule.ranges.add(xlrange)

    def valid_int(self, sheet, xlrange):
        """Set sheet range to int validation."""
        dvrule = DataValidation(type="whole")
        dvrule.errorTitle = self.XL_VALIDATION_ERROR_TITLE
        dvrule.error = 'Your entry is not a whole number.'
        sheet.add_data_validation(dvrule)
        dvrule.ranges.add(xlrange)

    def valid_dec(self, sheet, xlrange):
        """Set sheet range to int validation."""
        dvrule = DataValidation(type="decimal")
        dvrule.errorTitle = self.XL_VALIDATION_ERROR_TITLE
        dvrule.error = 'Your entry is not a decimal number.'
        sheet.add_data_validation(dvrule)
        dvrule.ranges.add(xlrange)

    def valid_str_len(self, sheet, xlrange, length):
        """Set sheet range to str length validation."""
        dvrule = DataValidation(
            type="textLength", operator="lessThanOrEqual", formula1=length)
        dvrule.errorTitle = self.XL_VALIDATION_ERROR_TITLE
        dvrule.error = 'Your entry exceeds {} characters.'.format(length)
        sheet.add_data_validation(dvrule)
        dvrule.ranges.add(xlrange)
