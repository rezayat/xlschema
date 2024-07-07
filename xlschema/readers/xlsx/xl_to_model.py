"""Converts xlsx files into xlschema models.

This basically means that they can then be converted into other formats.
"""

import sys

import openpyxl

from . import sheets
from .. import abstract
from ...config import Config

# ----------------------------------------------------------
# XLSX Conversion
# ----------------------------------------------------------


class ExcelToModel(abstract.SchemaReader):
    """Parses xlsx files into relational models."""

    def preprocess(self):
        """Runs before main process method for conversion.

        Place reader specific instance variable here
        """
        self.n_args = len(Config.METAFIELDS)
        self.workbook = openpyxl.load_workbook(self.uri, data_only=True)

    def _add_model_sheet(self, xlsheet, sheet_class):
        """Helper function to add model_sheet."""
        model_sheet = sheet_class(xlsheet, self.options)
        model_sheet.parse()
        self.schema.models.append(model_sheet.model)

    def process(self):
        """Main process for xl to sql conversion.

        1. Check 1st sheet for format to decide which way to go.

        2. Dispatch to respective subprocessing function.
        """
        has_properties = lambda sheet: sheet.cell(1, 1).value == 'app'
        id_in_cell = lambda sheet, row, col: sheet.cell(row, col).value == 'id'
        nargs = lambda col: '{}{}'.format(col, self.n_args)

        for name in self.workbook.sheetnames:
            sheet = self.workbook.get_sheet_by_name(name)

            if name == self.config.ENUMS_SHEET:
                enum_sheet = sheets.EnumSheet(sheet, self.options)
                enum_sheet.parse()
                self.schema.enums.update(enum_sheet.enums)
            else:
                # no data
                if id_in_cell(sheet, 3, 1):
                    self._add_model_sheet(sheet, sheets.NoDataSheet)

                # properties with data with left header
                elif has_properties(sheet):
                    self._add_model_sheet(sheet, sheets.PropertySheet)

                # data with left header
                elif id_in_cell(sheet, self.n_args, 2):
                    self._add_model_sheet(sheet, sheets.DataSheet)

                else:
                    self.log.critical('Cannot parse %s', self.uri)
                    sys.exit(1)

        self.post_process(self.workbook.sheetnames)
