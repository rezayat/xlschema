"""Splits xlsx files into several files along a column."""
import logging
import os

import openpyxl

from ..config import Config
from ..readers.xlsx import mixins
from .abstract import Plugin

# ----------------------------------------------------------
# XL File Splitter
# ----------------------------------------------------------


class XLSplitter(mixins.ToExcelMixin):
    """Splits xlsx files on a single column."""

    def __init__(self, path, target_sheet, group_on, keep_sheets=None,
                 outdir='splits', options=None):
        """Class constructor.

        :param path: path to xlsx file
        :type path: str

        :param target_sheet: name of sheet to be split
        :type target_sheet: str

        :param group_on: column to split on
        :type group_on: str

        :param keep_sheets: other sheet to be included in the outcome
        :type keep_sheets: List[str]

        :param outdir: name of subdirectory in output path which will contain
                       generated split sheets.
        :type outdir: str

        :param options: argparse options
        :type options: :py:class:`argparse.Namespace`
        """
        self.path = path
        self.target_sheet = target_sheet
        self.group_on = group_on
        self.keep_sheets = [target_sheet] + keep_sheets if keep_sheets else []

        self.parent_dir = os.path.dirname(path)
        self.output = os.path.join(options.output, outdir)
        self.tempfile = os.path.join(self.output, '_empty.xlsx')
        self.row_offset = len(Config.METAFIELDS)
        self.col_index = 1
        self.options = options

        # regular
        self.config = Config()
        self.log = logging.getLogger(self.__class__.__name__)

        # working instance vars
        self.groups = set()
        self.workbook = None
        self.data = []
        self.datasets = {}

    def process(self):
        """Process workbook."""
        self.xlsx_read()
        self.xlsx_write()
        self.cleanup()

    def xlsx_read(self):
        """Load and read workbook to be split."""
        self.log.debug('reading: %s', self.path)
        self.workbook = openpyxl.load_workbook(self.path)
        for name in self.workbook.get_sheet_names():
            if name == self.target_sheet:
                sheet = self.workbook.get_sheet_by_name(name)
                self.col_index = self._get_col_index(sheet)
                self.log.debug('column index: %s', self.col_index)
                self.read_sheet(sheet)

            if self.keep_sheets:
                if name not in self.keep_sheets:
                    sheet = self.workbook.get_sheet_by_name(name)
                    self.workbook.remove_sheet(sheet)

        if not os.path.exists(self.output):
            os.mkdir(self.output)
        self.log.debug('saving template: %s', self.tempfile)
        self.workbook.save(self.tempfile)

    def cleanup(self):
        """Drop tempfile."""
        os.remove(self.tempfile)

    def xlsx_write(self):
        """Write split workbook."""
        for name in sorted(self.datasets.keys()):
            outfile = os.path.join(self.output, '{}.xlsx'.format(name))
            self.workbook = openpyxl.load_workbook(self.tempfile)
            sheet = self.workbook.get_sheet_by_name(self.target_sheet)
            rows = self.datasets[name]
            for i, row in enumerate(rows, 1):
                for j, value in enumerate(row, 1):
                    cell = sheet.cell(row=i + self.row_offset, column=j)
                    cell.value = value
            self.log.debug('writing to file: %s', outfile)
            self.workbook.save(outfile)

    def read_sheet(self, sheet):
        """Read individual sheet."""
        self.log.debug('reading data from sheet: %s', self.target_sheet)
        # for row in sheet.iter_rows(row_offset=self.row_offset):
        for row in sheet.iter_rows(min_row=self.row_offset+1):
            self.groups.add(row[self.col_index].value)
            entry = []
            for cell in row:
                entry.append(cell.value)
                cell.value = None
            self.data.append(entry)

        groups = list(i for i in sorted(g for g in self.groups if g))
        self.log.debug('groups: %s', groups)
        while groups:
            target = groups.pop()
            self.datasets[target] = []
            for row in self.data:
                if target == row[self.col_index]:
                    self.datasets[target].append(row)

    def _get_col_index(self, sheet):
        """Get column index for the sheet."""
        # field_range = 'B{r}:CC{r}'.format(r=self.row_offset)
        # row = list(sheet.iter_rows(field_range))[0]
        row = list(sheet.iter_rows(
            min_row=self.row_offset, min_col=2, max_col=81))[0]
        print("row:", [c.value for c in row])
        index = 0
        for i, cell in enumerate(row):
            if cell.value == self.group_on:
                index = i + 1
        print("index:", index)
        return index

# ----------------------------------------------------------
# SplitterPlugin
# ----------------------------------------------------------


class SplitterPlugin(Plugin):
    """Splits xlsx sheets from a column."""

    name = 'splitter'
    subcommand = 'split_xlsx'
    requires = ['openpyxl']
    is_active = True

    @classmethod
    def setup_cmdline(cls, app):
        """Set up and register cmdline options for splitter plugin instance."""
        option = cls.register_plugin_subparser(app, cls)
        option('--keep_sheet', '-k', nargs='*', help="sheet(s) to keep")
        option('--group_on', '-g', help="group on field")
        option('--sheet', '-s', help="target sheet to split from")
        option('path', help="path of xlsx file to split")

    def execute(self, *args, **kwds):
        """Split xlsx sheets along a column."""
        self.log.debug('splitting %s', self.options.path)
        splitter = XLSplitter(
            path=self.options.path,
            target_sheet=self.options.sheet,
            group_on=self.options.group_on,
            keep_sheets=self.options.keep_sheet,
            options=self.options,
        )
        splitter.process()
        self.store['success'] = True
